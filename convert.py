"""
Converte un file .xlsx (listone fantacalcio) in listone.csv utilizzando openpyxl.
Uso:
    python convert.py <file.xlsx>              -> produce listone.csv
    python convert.py <file.xlsx> output.csv   -> produce output.csv
"""

import sys
import csv
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Errore: la libreria 'openpyxl' non è installata.")
    print("Installala con: pip install openpyxl")
    sys.exit(1)


def convert_xlsx_to_csv(input_path: str, output_path: str = 'listone.csv') -> None:
    """
    Legge il primo foglio del file xlsx e lo scrive come CSV con header.
    
    Args:
        input_path: Percorso del file .xlsx sorgente.
        output_path: Percorso del file .csv di destinazione (default: listone.csv).
    """
    path = Path(input_path)
    if not path.exists():
        print(f"Errore: file '{input_path}' non trovato.")
        sys.exit(1)
    if path.suffix.lower() != '.xlsx':
        print(f"Errore: il file deve avere estensione .xlsx (ricevuto: '{path.suffix}').")
        sys.exit(1)

    print(f"Apertura di '{input_path}'...")
    wb = load_workbook(input_path, read_only=True, data_only=True)
    ws = wb.active

    row_count = 0
    with open(output_path, 'w', newline='', encoding='utf-8') as out:
        writer = csv.writer(out)
        for row in ws.iter_rows(values_only=True):
            # Salta righe completamente vuote
            if all(cell is None for cell in row):
                continue
            writer.writerow([(cell if cell is not None else '') for cell in row])
            row_count += 1

    wb.close()

    # -1 per l'header
    print(f"Conversione completata: {row_count - 1} righe scritte in '{output_path}'.")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Uso: python convert.py <file.xlsx> [output.csv]")
        print("  Se non specificato, il file di output sarà 'listone.csv'.")
        sys.exit(1)

    xlsx_file = sys.argv[1]
    csv_file = sys.argv[2] if len(sys.argv) > 2 else 'listone.csv'
    convert_xlsx_to_csv(xlsx_file, csv_file)
